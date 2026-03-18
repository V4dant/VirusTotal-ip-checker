"""
VirusTotal IP Checker  (with API limit protection)
===================================================
HOW TO USE EVERY DAY:
  1. Paste your IPs into IP_LIST below
  2. Open Terminal and run:  python3 vt_ip_checker.py
  3. Excel is saved automatically

IF LIMIT RUNS OUT MID-RUN:
  - Script saves progress to  progress.json  automatically
  - Next day just run the script again — it resumes from where it stopped
  - When all IPs are done, progress.json is deleted automatically

FIRST TIME ONLY — install libraries:
  pip3 install requests openpyxl
"""

import time
import json
import os
import requests
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# ──────────────────────────────────────────────────────
#  1. PASTE YOUR API KEY
# ──────────────────────────────────────────────────────

VT_API_KEY = "c3cf9f5b37517a9771c510d37a5de6adbe2fae7a6ed5bdfe9baa654638ecee40"

# ──────────────────────────────────────────────────────
#  2. PASTE YOUR IPs HERE  (limit : 500)
# ──────────────────────────────────────────────────────
#paste your ips here, one per line, up to 500 total
IP_LIST = """
110.42.110.67
""".strip()

# ──────────────────────────────────────────────────────
#  SETTINGS
# ──────────────────────────────────────────────────────

REQUESTS_PER_MINUTE = 20000          # Free API = 4/min, 500/day
PROGRESS_FILE       = "progress.json"
OUTPUT_FILE         = f"IP_Threat_Analysis_{date.today()}.xlsx"


# ──────────────────────────────────────────────────────
#  PROGRESS  (save & resume)
# ──────────────────────────────────────────────────────

def load_progress(ips):
    """
    If progress.json exists AND has the same IP list,
    reload previous results so we skip already-checked IPs.
    Otherwise start fresh.
    """
    if not os.path.exists(PROGRESS_FILE):
        return {}

    try:
        with open(PROGRESS_FILE, "r") as f:
            saved = json.load(f)

        if saved.get("ip_list") != ips:
            print("  Detected a NEW IP list — starting fresh.\n")
            return {}

        results = saved.get("results", {})
        print(f"  Resuming from previous run — {len(results)} IPs already done.\n")
        return results

    except Exception:
        return {}


def save_progress(ips, results):
    """Save current results to disk so we can resume if limit hits."""
    with open(PROGRESS_FILE, "w") as f:
        json.dump({"ip_list": ips, "results": results}, f)


def clear_progress():
    """Delete progress file once everything is complete."""
    if os.path.exists(PROGRESS_FILE):
        os.remove(PROGRESS_FILE)


# ──────────────────────────────────────────────────────
#  VIRUSTOTAL LOOKUP
# ──────────────────────────────────────────────────────

def check_ip(ip):
    """
    Returns:
      (malicious_count, total_engines)  — normal result
      (None, None)                      — IP not in VT
      "LIMIT"                           — daily quota exceeded
      "ERROR"                           — network or other error
    """
    url     = f"https://www.virustotal.com/api/v3/ip_addresses/{ip}"
    headers = {"x-apikey": VT_API_KEY}

    try:
        response = requests.get(url, headers=headers, timeout=15)

        if response.status_code == 200:
            stats     = response.json()["data"]["attributes"]["last_analysis_stats"]
            malicious = stats.get("malicious", 0)
            total     = sum(stats.values())
            return malicious, total

        elif response.status_code == 404:
            return None, None  # IP not in VT database

        elif response.status_code == 204:
            # 204 = quota exceeded (free API daily limit hit)
            return "LIMIT", None

        elif response.status_code == 429:
            # 429 = per-minute rate limit — wait and retry once
            print("  Per-minute rate limit — waiting 60s...")
            time.sleep(60)
            return check_ip(ip)

        else:
            print(f"  HTTP {response.status_code}")
            return "ERROR", None

    except requests.exceptions.Timeout:
        print("  Timeout — skipping")
        return "ERROR", None

    except Exception as e:
        print(f"  Error: {e}")
        return "ERROR", None


# ──────────────────────────────────────────────────────
#  EXCEL BUILDER
# ──────────────────────────────────────────────────────

def build_excel(ips, results_map):
    """
    ips         = original ordered list
    results_map = dict { ip: (malicious, total) or None or "LIMIT" or "ERROR" }

    Columns:
      A — Source Address
      B — Malicious / Clean / Unknown / Pending
      C — VT Rating  e.g. 5//94  (blank if clean/unknown/pending)
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "IP Threat Analysis"

    # Header
    hdr_font = Font(name="Calibri", bold=True, color="000000", size=12)

    for col, title in enumerate(["Source Address", "Status", "VT Rating"], 1):
        c = ws.cell(row=1, column=col, value=title)
        c.font = hdr_font

    # Row styles — black Calibri 12, no colours
    plain_font = Font(name="Calibri", color="000000", size=12)
    center     = Alignment(horizontal="center")

    for row_num, ip in enumerate(ips, start=2):
        result = results_map.get(ip, "PENDING")

        if result == "PENDING" or result == "LIMIT":
            status = "Pending (limit hit)"
            rating = ""
        elif result == "ERROR" or result == (None, None):
            status = "Unknown"
            rating = ""
        else:
            malicious, total = result
            if malicious == 0:
                status = "Clean"
                rating = ""
            else:
                status = "Malicious"
                rating = f"{malicious}//{total}"

        a = ws.cell(row=row_num, column=1, value=ip)
        a.font = plain_font

        b = ws.cell(row=row_num, column=2, value=status)
        b.font = plain_font; b.alignment = center

        c = ws.cell(row=row_num, column=3, value=rating)
        c.font = plain_font; c.alignment = center

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 14
    ws.freeze_panes = "A2"

    # Summary sheet
    ws2 = wb.create_sheet("Summary")
    mal     = sum(1 for ip in ips if isinstance(results_map.get(ip), tuple) and results_map[ip][0] > 0)
    cln     = sum(1 for ip in ips if isinstance(results_map.get(ip), tuple) and results_map[ip][0] == 0)
    unk     = sum(1 for ip in ips if results_map.get(ip) in ["ERROR", (None, None)])
    pending = sum(1 for ip in ips if results_map.get(ip) in ["PENDING", "LIMIT", None])

    ws2["A1"] = "Summary"
    ws2["A1"].font = Font(name="Arial", bold=True, size=14)

    summary_rows = [
        ("Total IPs",     len(ips),   "000000"),
        ("Malicious",     mal,        "C0392B"),
        ("Clean",         cln,        "1E8449"),
        ("Unknown",       unk,        "888888"),
        ("Pending",       pending,    "B7860B"),
    ]
    for i, (label, val, color) in enumerate(summary_rows, start=3):
        ws2.cell(row=i, column=1, value=label).font = Font(name="Arial", bold=True, color=color, size=11)
        ws2.cell(row=i, column=2, value=val).font   = Font(name="Arial", color=color, size=11)

    ws2.column_dimensions["A"].width = 16
    ws2.column_dimensions["B"].width = 10

    wb.save(OUTPUT_FILE)


# ──────────────────────────────────────────────────────
#  MAIN
# ──────────────────────────────────────────────────────

def main():
    if VT_API_KEY == "PASTE_YOUR_API_KEY_HERE":
        print("\nERROR: Set your VT_API_KEY in the script first.\n")
        return

    ips = [line.strip() for line in IP_LIST.splitlines() if line.strip()]
    if not ips:
        print("\nERROR: No IPs found in IP_LIST.\n")
        return

    delay = 60 / REQUESTS_PER_MINUTE

    print(f"\nVirusTotal IP Checker")
    print(f"{'─'*44}")
    print(f"  Total IPs    : {len(ips)}")
    print(f"  Rate limit   : {REQUESTS_PER_MINUTE} req/min  |  500 req/day (free)")
    print(f"  Est. time    : ~{round(len(ips) * delay / 60, 1)} minutes")
    print(f"{'─'*44}\n")

    # Load any previous progress for this exact IP list
    results_map = load_progress(ips)

    limit_hit   = False
    checked_now = 0

    for idx, ip in enumerate(ips, start=1):

        # Already checked in a previous run — skip
        if ip in results_map:
            print(f"[{idx:>3}/{len(ips)}]  {ip:<20}  (already done — skipping)")
            continue

        # If limit hit earlier in THIS run, mark remaining as pending and stop
        if limit_hit:
            results_map[ip] = "LIMIT"
            continue

        print(f"[{idx:>3}/{len(ips)}]  {ip:<20}", end="  ", flush=True)

        result = check_ip(ip)

        # ── Quota exceeded ──────────────────────────────
        if result == ("LIMIT", None):
            print("DAILY LIMIT REACHED — saving progress and stopping.")
            results_map[ip] = "LIMIT"
            limit_hit = True
            save_progress(ips, results_map)
            continue

        # ── Error ───────────────────────────────────────
        elif result == ("ERROR", None):
            print("error — marked as Unknown")
            results_map[ip] = "ERROR"

        # ── Not in VT ───────────────────────────────────
        elif result == (None, None):
            print("not found in VT")
            results_map[ip] = (None, None)

        # ── Normal result ───────────────────────────────
        else:
            malicious, total = result
            results_map[ip]  = (malicious, total)
            if malicious == 0:
                print("Clean")
            else:
                print(f"MALICIOUS  {malicious}//{total}")

        checked_now += 1

        # Save progress after every IP so nothing is lost
        save_progress(ips, results_map)

        # Wait between requests (skip after last IP)
        if idx < len(ips) and not limit_hit:
            time.sleep(delay)

    # ── Build Excel (always — even if partial) ──────────
    print(f"\n{'─'*44}")

    if limit_hit:
        remaining = sum(1 for ip in ips if results_map.get(ip) in ["LIMIT", None])
        print(f"  Limit hit! {remaining} IPs still pending.")
        print(f"  Run the script again tomorrow to finish.\n")
    else:
        print(f"  All IPs checked!")
        clear_progress()  # clean up — no longer needed

    print(f"  Building Excel → {OUTPUT_FILE}")
    build_excel(ips, results_map)

    mal     = sum(1 for ip in ips if isinstance(results_map.get(ip), tuple) and results_map[ip][0] is not None and results_map[ip][0] > 0)
    cln     = sum(1 for ip in ips if isinstance(results_map.get(ip), tuple) and results_map[ip][0] == 0)
    pending = sum(1 for ip in ips if results_map.get(ip) in ["LIMIT", None])

    print(f"\n  Done!")
    print(f"  Malicious : {mal}")
    print(f"  Clean     : {cln}")
    print(f"  Pending   : {pending}")
    print(f"  File      : {OUTPUT_FILE}\n")


if __name__ == "__main__":
    main()
